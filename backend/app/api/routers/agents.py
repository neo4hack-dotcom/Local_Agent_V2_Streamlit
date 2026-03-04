from __future__ import annotations

import json
from datetime import date, datetime, time, timezone
from typing import Any, Literal
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from app.api.deps import get_agents_repo, get_settings_repo
from app.core.agent_executor import AgentExecutor
from app.core.agent_templates import list_agent_templates
from app.core.database_routing import resolve_database_for_agent
from app.core.multi_agent_manager import MultiAgentManager
from app.core.models import (
    AgentCatalog,
    AgentConfigExport,
    AgentConfigImportRequest,
    AgentConfigImportResponse,
    AgentAuditEntry,
    AgentConfig,
    AgentConfigCreate,
    AgentConfigUpdate,
    AgentTemplate,
    AgentRunRequest,
    AgentRunResponse,
    ConversationTurn,
    ManagerRunRequest,
    ManagerRunResponse,
    AppSettings,
    agent_requires_database,
)
from app.core.settings import DATA_DIR
from app.core.storage import JSONRepository
from app.core.webhook_dispatcher import WebhookDispatcher

router = APIRouter(prefix="/agents", tags=["agents"])
MAX_AGENT_AUDIT_VERSIONS = 5
MAX_CONVERSATION_MEMORY_TURNS = 12
MAX_CONVERSATION_MEMORY_CHARS = 1200
MAX_EXCEL_TEXT_CHARS = 32000
MAX_EXCEL_PREVIEW_CHARS = 1200


def _find_agent(catalog: AgentCatalog, agent_id: str) -> AgentConfig:
    for agent in catalog.agents:
        if agent.id == agent_id:
            return agent
    raise HTTPException(status_code=404, detail="Agent not found.")


def _find_agent_index(catalog: AgentCatalog, agent_id: str) -> int:
    for index, agent in enumerate(catalog.agents):
        if agent.id == agent_id:
            return index
    raise HTTPException(status_code=404, detail="Agent not found.")


def _append_agent_audit(
    catalog: AgentCatalog,
    agent: AgentConfig,
    reason: Literal["update", "restore"],
) -> None:
    history = catalog.audit_history.setdefault(agent.id, [])
    history.insert(
        0,
        AgentAuditEntry(
            version_id=uuid4().hex,
            created_at=datetime.now(timezone.utc).isoformat(),
            reason=reason,
            snapshot=agent.model_copy(deep=True),
        ),
    )
    if len(history) > MAX_AGENT_AUDIT_VERSIONS:
        del history[MAX_AGENT_AUDIT_VERSIONS:]


def _count_audit_entries(history: dict[str, list[AgentAuditEntry]]) -> int:
    return sum(len(entries) for entries in history.values())


def _validate_unique_agent_ids(agents: list[AgentConfig]) -> None:
    seen: set[str] = set()
    duplicated: set[str] = set()
    for agent in agents:
        if agent.id in seen:
            duplicated.add(agent.id)
        seen.add(agent.id)
    if duplicated:
        duplicated_text = ", ".join(sorted(duplicated))
        raise HTTPException(
            status_code=400,
            detail=f"Import contains duplicated agent IDs: {duplicated_text}",
        )


def _truncate_text(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return value[: max(0, limit - 3)].rstrip() + "..."


def _normalized_conversation_history(turns: list[ConversationTurn]) -> list[ConversationTurn]:
    normalized: list[ConversationTurn] = []
    for turn in turns[-MAX_CONVERSATION_MEMORY_TURNS:]:
        content = str(turn.content).strip()
        if not content:
            continue
        normalized.append(
            ConversationTurn(
                role=turn.role,
                content=_truncate_text(content, MAX_CONVERSATION_MEMORY_CHARS),
            )
        )
    return normalized


def _contextual_question(question: str, turns: list[ConversationTurn]) -> str:
    base_question = str(question).strip()
    if not turns:
        return base_question

    memory_lines = ["Conversation memory (oldest to newest):"]
    for turn in turns:
        memory_lines.append(f"- {turn.role}: {turn.content}")

    return (
        "\n".join(memory_lines)
        + "\n\nCurrent request:\n"
        + base_question
        + "\n\nInstruction: Use memory only if relevant and prioritize the current request."
    )


def _json_default(value: Any) -> str:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _safe_json_text(value: Any, limit: int = MAX_EXCEL_TEXT_CHARS) -> str:
    try:
        text = json.dumps(value, ensure_ascii=False, default=_json_default)
    except Exception:  # noqa: BLE001
        text = str(value)
    return _truncate_text(text, limit)


def _excel_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, datetime, date, time)):
        return value
    if isinstance(value, str):
        return _truncate_text(value, MAX_EXCEL_TEXT_CHARS)
    if isinstance(value, (dict, list, tuple, set)):
        return _safe_json_text(value, limit=MAX_EXCEL_TEXT_CHARS)
    return _truncate_text(str(value), MAX_EXCEL_TEXT_CHARS)


def _call_event_key(event: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(event.get("step", "")),
        str(event.get("call_index", "")),
        str(event.get("agent_id", "")),
    )


def _export_manager_intermediate_results_to_excel(
    *,
    question: str,
    timeline: list[dict[str, Any]],
) -> tuple[str | None, str | None]:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError:
        return (
            None,
            "Excel export requires 'openpyxl'. Install backend dependencies and retry.",
        )

    try:
        export_dir = DATA_DIR / "exports" / "manager_runs"
        export_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_path = export_dir / f"manager_intermediate_{timestamp}_{uuid4().hex[:8]}.xlsx"

        workbook = Workbook()
        summary_ws = workbook.active
        summary_ws.title = "run_summary"

        final_event = next(
            (event for event in reversed(timeline) if event.get("type") == "manager_final"),
            None,
        )

        summary_ws.append(["field", "value"])
        summary_rows: list[tuple[str, Any]] = [
            ("generated_at_utc", datetime.now(timezone.utc).isoformat()),
            ("question", question),
            ("total_events", len(timeline)),
            ("final_status", final_event.get("status") if isinstance(final_event, dict) else ""),
            ("final_answer", final_event.get("answer") if isinstance(final_event, dict) else ""),
            (
                "final_missing_information",
                final_event.get("missing_information") if isinstance(final_event, dict) else "",
            ),
            ("final_steps", final_event.get("steps") if isinstance(final_event, dict) else ""),
            (
                "final_agent_calls",
                final_event.get("agent_calls") if isinstance(final_event, dict) else "",
            ),
        ]
        for key, value in summary_rows:
            summary_ws.append([key, _excel_cell_value(value)])

        events_ws = workbook.create_sheet("timeline_events")
        events_ws.append(
            [
                "index",
                "ts",
                "type",
                "step",
                "call_index",
                "agent_id",
                "agent_name",
                "agent_type",
                "status",
                "question",
                "rationale",
                "error",
                "row_count",
                "database_id",
                "database_name",
                "missing_information",
                "answer_preview",
                "sql_preview",
                "details_json",
                "rows_preview_json",
            ]
        )
        for index, event in enumerate(timeline, start=1):
            if not isinstance(event, dict):
                continue
            events_ws.append(
                [
                    index,
                    _excel_cell_value(event.get("ts")),
                    _excel_cell_value(event.get("type")),
                    _excel_cell_value(event.get("step")),
                    _excel_cell_value(event.get("call_index")),
                    _excel_cell_value(event.get("agent_id")),
                    _excel_cell_value(event.get("agent_name")),
                    _excel_cell_value(event.get("agent_type")),
                    _excel_cell_value(event.get("status")),
                    _excel_cell_value(event.get("question")),
                    _excel_cell_value(event.get("rationale")),
                    _excel_cell_value(event.get("error")),
                    _excel_cell_value(event.get("row_count")),
                    _excel_cell_value(event.get("database_id")),
                    _excel_cell_value(event.get("database_name")),
                    _excel_cell_value(event.get("missing_information")),
                    _truncate_text(str(event.get("answer", "")), MAX_EXCEL_PREVIEW_CHARS),
                    _truncate_text(str(event.get("sql", "")), MAX_EXCEL_PREVIEW_CHARS),
                    _safe_json_text(event.get("details", {}), limit=MAX_EXCEL_TEXT_CHARS),
                    _safe_json_text(event.get("rows_preview", []), limit=MAX_EXCEL_TEXT_CHARS),
                ]
            )

        calls_ws = workbook.create_sheet("agent_calls")
        calls_ws.append(
            [
                "step",
                "call_index",
                "agent_id",
                "agent_name",
                "agent_type",
                "question",
                "status",
                "row_count",
                "error",
                "answer_preview",
                "sql_preview",
                "database_id",
                "database_name",
                "details_json",
            ]
        )
        started_calls: dict[tuple[str, str, str], dict[str, Any]] = {}
        for event in timeline:
            if not isinstance(event, dict):
                continue
            event_type = str(event.get("type", ""))
            if event_type == "agent_call_started":
                started_calls[_call_event_key(event)] = event
                continue
            if event_type not in {"agent_call_completed", "agent_call_failed"}:
                continue
            key = _call_event_key(event)
            started = started_calls.pop(key, {})
            calls_ws.append(
                [
                    _excel_cell_value(event.get("step")),
                    _excel_cell_value(event.get("call_index")),
                    _excel_cell_value(event.get("agent_id")),
                    _excel_cell_value(event.get("agent_name")),
                    _excel_cell_value(event.get("agent_type")),
                    _excel_cell_value(
                        started.get("question")
                        if isinstance(started, dict)
                        else event.get("question")
                    ),
                    "success" if event_type == "agent_call_completed" else "failed",
                    _excel_cell_value(event.get("row_count")),
                    _excel_cell_value(event.get("error")),
                    _truncate_text(str(event.get("answer", "")), MAX_EXCEL_PREVIEW_CHARS),
                    _truncate_text(str(event.get("sql", "")), MAX_EXCEL_PREVIEW_CHARS),
                    _excel_cell_value(
                        started.get("database_id")
                        if isinstance(started, dict)
                        else event.get("database_id")
                    ),
                    _excel_cell_value(
                        started.get("database_name")
                        if isinstance(started, dict)
                        else event.get("database_name")
                    ),
                    _safe_json_text(event.get("details", {}), limit=MAX_EXCEL_TEXT_CHARS),
                ]
            )

        raw_ws = workbook.create_sheet("raw_events")
        raw_ws.append(["index", "event_json"])
        for index, event in enumerate(timeline, start=1):
            raw_ws.append([index, _safe_json_text(event, limit=MAX_EXCEL_TEXT_CHARS)])

        workbook.save(file_path)
        return str(file_path), None
    except Exception as exc:  # noqa: BLE001
        return None, str(exc)


@router.get("", response_model=list[AgentConfig])
def list_agents(
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
) -> list[AgentConfig]:
    return agents_repo.load().agents


@router.get("/templates", response_model=list[AgentTemplate])
def get_agent_templates() -> list[AgentTemplate]:
    return list_agent_templates()


@router.get("/export", response_model=AgentConfigExport)
def export_agents_config(
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
) -> AgentConfigExport:
    catalog = agents_repo.load()
    return AgentConfigExport(
        exported_at=datetime.now(timezone.utc).isoformat(),
        agents=[agent.model_copy(deep=True) for agent in catalog.agents],
        audit_history={
            agent_id: [entry.model_copy(deep=True) for entry in entries]
            for agent_id, entries in catalog.audit_history.items()
        },
    )


@router.post("/import", response_model=AgentConfigImportResponse)
def import_agents_config(
    payload: AgentConfigImportRequest,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
) -> AgentConfigImportResponse:
    catalog = agents_repo.load()
    imported_agents = [agent.model_copy(deep=True) for agent in payload.payload.agents]
    _validate_unique_agent_ids(imported_agents)

    imported_history = {
        agent_id: [
            entry.model_copy(deep=True) for entry in entries[:MAX_AGENT_AUDIT_VERSIONS]
        ]
        for agent_id, entries in payload.payload.audit_history.items()
    }
    imported_ids = {agent.id for agent in imported_agents}

    if payload.mode == "replace":
        catalog.agents = imported_agents
        if payload.preserve_audit_history:
            catalog.audit_history = {
                agent_id: imported_history.get(agent_id, [])
                for agent_id in imported_ids
            }
        else:
            catalog.audit_history = {}
        agents_repo.save(catalog)
        return AgentConfigImportResponse(
            mode=payload.mode,
            imported_agents=len(imported_agents),
            created_agents=0,
            updated_agents=0,
            replaced_agents=len(imported_agents),
            preserved_audit_entries=_count_audit_entries(catalog.audit_history),
        )

    existing_by_id: dict[str, AgentConfig] = {agent.id: agent for agent in catalog.agents}
    created_agents = 0
    updated_agents = 0
    for agent in imported_agents:
        if agent.id in existing_by_id:
            updated_agents += 1
        else:
            created_agents += 1
        existing_by_id[agent.id] = agent

    merged_order: list[str] = [agent.id for agent in catalog.agents]
    for agent in imported_agents:
        if agent.id not in merged_order:
            merged_order.append(agent.id)
    catalog.agents = [existing_by_id[agent_id] for agent_id in merged_order if agent_id in existing_by_id]

    if payload.preserve_audit_history:
        for agent_id, entries in imported_history.items():
            if agent_id in existing_by_id:
                catalog.audit_history[agent_id] = entries

    agents_repo.save(catalog)
    return AgentConfigImportResponse(
        mode=payload.mode,
        imported_agents=len(imported_agents),
        created_agents=created_agents,
        updated_agents=updated_agents,
        replaced_agents=0,
        preserved_audit_entries=_count_audit_entries(catalog.audit_history),
    )


@router.post("", response_model=AgentConfig, status_code=201)
def create_agent(
    payload: AgentConfigCreate,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
) -> AgentConfig:
    catalog = agents_repo.load()
    agent = AgentConfig(id=uuid4().hex, **payload.model_dump())
    catalog.agents.append(agent)
    agents_repo.save(catalog)
    return agent


@router.put("/{agent_id}", response_model=AgentConfig)
def update_agent(
    agent_id: str,
    payload: AgentConfigUpdate,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
) -> AgentConfig:
    catalog = agents_repo.load()
    index = _find_agent_index(catalog, agent_id)
    previous = catalog.agents[index]
    _append_agent_audit(catalog, previous, reason="update")
    updated = AgentConfig(id=agent_id, **payload.model_dump())
    catalog.agents[index] = updated
    agents_repo.save(catalog)
    return updated


@router.get("/{agent_id}/audit", response_model=list[AgentAuditEntry])
def list_agent_audit(
    agent_id: str,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
) -> list[AgentAuditEntry]:
    catalog = agents_repo.load()
    _find_agent(catalog, agent_id)
    return catalog.audit_history.get(agent_id, [])


@router.post("/{agent_id}/audit/{version_id}/restore", response_model=AgentConfig)
def restore_agent_version(
    agent_id: str,
    version_id: str,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
) -> AgentConfig:
    catalog = agents_repo.load()
    index = _find_agent_index(catalog, agent_id)
    current = catalog.agents[index]
    history = catalog.audit_history.get(agent_id, [])

    selected_version = next(
        (entry for entry in history if entry.version_id == version_id),
        None,
    )
    if not selected_version:
        raise HTTPException(status_code=404, detail="Agent version not found.")

    _append_agent_audit(catalog, current, reason="restore")
    restored = selected_version.snapshot.model_copy(update={"id": agent_id}, deep=True)
    catalog.agents[index] = restored
    agents_repo.save(catalog)
    return restored


@router.delete("/{agent_id}", status_code=204)
def delete_agent(
    agent_id: str,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
) -> None:
    catalog = agents_repo.load()
    catalog.agents = [agent for agent in catalog.agents if agent.id != agent_id]
    catalog.audit_history.pop(agent_id, None)
    agents_repo.save(catalog)


@router.post("/{agent_id}/run", response_model=AgentRunResponse)
def run_agent(
    agent_id: str,
    payload: AgentRunRequest,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
    settings_repo: JSONRepository[AppSettings] = Depends(get_settings_repo),
) -> AgentRunResponse:
    catalog = agents_repo.load()
    settings = settings_repo.load()

    agent = _find_agent(catalog, agent_id)
    if not agent.enabled:
        raise HTTPException(status_code=400, detail="This agent is disabled.")

    try:
        database = resolve_database_for_agent(
            agent=agent,
            databases=settings.databases,
            active_database_id=settings.active_database_id,
            requested_database_id=payload.database_id,
            required=agent_requires_database(agent.agent_type),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    memory_turns = _normalized_conversation_history(payload.conversation_history)
    contextual_question = _contextual_question(payload.question, memory_turns)

    executor = AgentExecutor(settings.llm)
    try:
        output = executor.execute(agent=agent, question=contextual_question, database=database)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    details = output.get("details", {})
    if not isinstance(details, dict):
        details = {}
    details = {
        **details,
        "memory_turns_used": len(memory_turns),
    }

    return AgentRunResponse(
        agent_id=agent.id,
        database_id=database.id if database else None,
        sql=output["sql"],
        rows=output["rows"],
        answer=output["answer"],
        details=details,
    )


@router.post("/manager/run", response_model=ManagerRunResponse)
def run_with_manager(
    payload: ManagerRunRequest,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
    settings_repo: JSONRepository[AppSettings] = Depends(get_settings_repo),
) -> ManagerRunResponse:
    catalog = agents_repo.load()
    settings = settings_repo.load()
    memory_turns = _normalized_conversation_history(payload.conversation_history)
    manager = MultiAgentManager(
        llm_config=settings.llm,
        agents=catalog.agents,
        databases=settings.databases,
        active_database_id=settings.active_database_id,
        requested_database_id=payload.database_id,
        conversation_memory=memory_turns,
    )

    timeline = list(manager.run_stream(payload))

    final_event = next(
        (event for event in reversed(timeline) if event.get("type") == "manager_final"),
        None,
    )

    if not final_event:
        raise HTTPException(status_code=500, detail="Manager execution ended unexpectedly.")

    if payload.export_intermediate_results_to_excel:
        export_path, export_error = _export_manager_intermediate_results_to_excel(
            question=payload.question,
            timeline=timeline,
        )
        if export_path:
            final_event["intermediate_results_excel_path"] = export_path
        if export_error:
            final_event["intermediate_results_excel_error"] = export_error

    if settings.webhook.enabled and settings.webhook.url.strip():
        dispatcher = WebhookDispatcher(settings.webhook)
        run_id = uuid4().hex
        run_context = {
            "channel": "manager_sync",
            "question": payload.question,
            "database_id": payload.database_id,
            "max_steps": payload.max_steps,
            "max_agent_calls": payload.max_agent_calls,
            "memory_turns": len(memory_turns),
            "export_intermediate_results_to_excel": payload.export_intermediate_results_to_excel,
        }
        for sequence, event in enumerate(timeline, start=1):
            dispatcher.send_manager_event(
                run_id=run_id,
                run_context=run_context,
                event=event,
                sequence=sequence,
                timeline=timeline if event.get("type") == "manager_final" else None,
            )

    return ManagerRunResponse(
        status=final_event["status"],
        answer=final_event["answer"],
        manager_summary=final_event.get("manager_summary"),
        judge_verdict=final_event.get("judge_verdict"),
        judge_confidence=final_event.get("judge_confidence"),
        judge_rationale=final_event.get("judge_rationale"),
        judge_checks_passed=final_event.get("judge_checks_passed", []),
        judge_checks_failed=final_event.get("judge_checks_failed", []),
        judge_recommendations=final_event.get("judge_recommendations", []),
        missing_information=final_event.get("missing_information"),
        intermediate_results_excel_path=final_event.get("intermediate_results_excel_path"),
        intermediate_results_excel_error=final_event.get("intermediate_results_excel_error"),
        steps=final_event["steps"],
        agent_calls=final_event["agent_calls"],
        timeline=timeline,
    )


@router.post("/manager/run/stream")
def run_with_manager_stream(
    payload: ManagerRunRequest,
    agents_repo: JSONRepository[AgentCatalog] = Depends(get_agents_repo),
    settings_repo: JSONRepository[AppSettings] = Depends(get_settings_repo),
) -> StreamingResponse:
    catalog = agents_repo.load()
    settings = settings_repo.load()
    memory_turns = _normalized_conversation_history(payload.conversation_history)
    webhook_dispatcher = WebhookDispatcher(settings.webhook)
    webhook_enabled = settings.webhook.enabled and settings.webhook.url.strip() != ""
    webhook_run_id = uuid4().hex
    webhook_run_context = {
        "channel": "manager_stream",
        "question": payload.question,
        "database_id": payload.database_id,
        "max_steps": payload.max_steps,
        "max_agent_calls": payload.max_agent_calls,
        "memory_turns": len(memory_turns),
        "export_intermediate_results_to_excel": payload.export_intermediate_results_to_excel,
    }

    def iter_events():
        manager = MultiAgentManager(
            llm_config=settings.llm,
            agents=catalog.agents,
            databases=settings.databases,
            active_database_id=settings.active_database_id,
            requested_database_id=payload.database_id,
            conversation_memory=memory_turns,
        )
        timeline_for_webhook: list[dict] = []
        sequence = 0
        try:
            for event in manager.run_stream(payload):
                sequence += 1
                timeline_for_webhook.append(event)
                if (
                    payload.export_intermediate_results_to_excel
                    and event.get("type") == "manager_final"
                ):
                    export_path, export_error = _export_manager_intermediate_results_to_excel(
                        question=payload.question,
                        timeline=timeline_for_webhook,
                    )
                    if export_path:
                        event["intermediate_results_excel_path"] = export_path
                    if export_error:
                        event["intermediate_results_excel_error"] = export_error
                if webhook_enabled:
                    webhook_dispatcher.send_manager_event(
                        run_id=webhook_run_id,
                        run_context=webhook_run_context,
                        event=event,
                        sequence=sequence,
                        timeline=timeline_for_webhook if event.get("type") == "manager_final" else None,
                    )
                yield json.dumps(event, ensure_ascii=False) + "\n"
        except Exception as exc:  # noqa: BLE001
            fallback = {
                "type": "manager_final",
                "status": "blocked",
                "answer": "The manager stopped because an unexpected error occurred.",
                "manager_summary": (
                    "What worked:\n"
                    "- none\n"
                    "What did not work:\n"
                    f"- Unexpected backend error: {str(exc)}\n"
                    "Agents called and purpose:\n"
                    "- none"
                ),
                "judge_verdict": "fail",
                "judge_confidence": 0,
                "judge_rationale": "Sanity check unavailable due to unexpected backend error.",
                "judge_checks_passed": [],
                "judge_checks_failed": ["Unexpected backend error interrupted orchestration."],
                "judge_recommendations": ["Inspect backend logs and rerun the request."],
                "missing_information": str(exc),
                "intermediate_results_excel_path": None,
                "intermediate_results_excel_error": None,
                "steps": 0,
                "agent_calls": 0,
            }
            if payload.export_intermediate_results_to_excel:
                export_path, export_error = _export_manager_intermediate_results_to_excel(
                    question=payload.question,
                    timeline=[*timeline_for_webhook, fallback],
                )
                fallback["intermediate_results_excel_path"] = export_path
                fallback["intermediate_results_excel_error"] = export_error
            sequence += 1
            timeline_for_webhook.append(fallback)
            if webhook_enabled:
                webhook_dispatcher.send_manager_event(
                    run_id=webhook_run_id,
                    run_context=webhook_run_context,
                    event=fallback,
                    sequence=sequence,
                    timeline=timeline_for_webhook,
                )
            yield json.dumps(fallback, ensure_ascii=False) + "\n"

    return StreamingResponse(iter_events(), media_type="application/x-ndjson")
